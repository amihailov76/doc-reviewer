"""
Роутер снимков оценки.

Итоговые снимки привязаны к продуктовым группам (group_id).
Промежуточные снимки (is_partial=True) group_id не требуют — хранятся на уровне документа.
document_id и document_filename хранятся как метаданные в обоих типах.

Эндпойнты:
  GET    /api/snapshots/document/{id}/check-evaluated  — есть ли оценки
  GET    /api/snapshots/document/{id}/partial          — список промежуточных снимков документа
  POST   /api/snapshots/document/{id}                  — создать снимок (итоговый или промежуточный)
  POST   /api/snapshots/merge                          — объединить промежуточные → итоговый снимок
  DELETE /api/snapshots/{id}                           — удалить снимок
  GET    /api/snapshots/compare                        — сравнить два снимка / снимок с текущим
  GET    /api/snapshots/document/{id}/export           — скачать XLS
"""

import io
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import List, Optional

from backend.database import get_db, Document, Instruction, Evaluation, Snapshot, ProductGroup

router = APIRouter(prefix="/api/snapshots", tags=["snapshots"])

COLOR_RANK = {"green": 0, "yellow": 1, "orange": 2, "red": 3}
COLOR_LABEL = {"green": "Хорошо", "yellow": "Замечания", "orange": "Проблемы", "red": "Критично"}


class CreateSnapshotRequest(BaseModel):
    name: str
    role: str = "current"
    group_id: Optional[int] = None   # обязателен для итоговых; для промежуточных — не нужен
    is_partial: bool = False


class MergeSnapshotsRequest(BaseModel):
    snapshot_ids: List[int]
    name: str
    group_id: Optional[int] = None   # опционально: куда поместить итоговый снимок
    role: str = "current"


def _serialize_snapshot(s: Snapshot) -> dict:
    return {
        "id": s.id,
        "group_id": s.group_id,
        "document_id": s.document_id,
        "document_filename": s.document_filename,
        "name": s.name,
        "role": s.role,
        "is_partial": bool(s.is_partial),
        "created_at": s.created_at.isoformat(),
        "data": s.data,
    }


def _compute_summary(evaluated: list) -> dict:
    summary = {"green": 0, "yellow": 0, "orange": 0, "red": 0, "total": 0}
    for instr in evaluated:
        color = instr.evaluation.color
        if color in summary:
            summary[color] += 1
        summary["total"] += 1
    return summary


def _compute_summary_from_sections(sections: list) -> dict:
    """Пересчитывает сводку из списка секций (используется при слиянии снимков)."""
    summary = {"green": 0, "yellow": 0, "orange": 0, "red": 0, "total": 0}
    for sec in sections:
        color = sec.get("color")
        if color in summary:
            summary[color] += 1
        summary["total"] += 1
    return summary


_COLOR_POINTS = {"green": 3, "yellow": 2, "orange": 1, "red": 0}
_GRADE_THRESHOLDS = [(85, "A", "Хорошо"), (65, "B", "Есть замечания"), (40, "C", "Требует доработки"), (0, "D", "Критично")]


def _compute_integral_from_instructions(evaluated: list) -> Optional[dict]:
    """Вычисляет интегральную оценку из списка оценённых инструкций (для снимка)."""
    if not evaluated:
        return None
    total_points = 0
    violation_counts: dict = {}
    for instr in evaluated:
        ev = instr.evaluation
        color = ev.color or "red"
        total_points += _COLOR_POINTS.get(color, 0)
        for crit_id, result in (ev.criteria_results or {}).items():
            if isinstance(result, dict) and result.get("result") == "error":
                label = result.get("label") or crit_id
                if crit_id not in violation_counts:
                    violation_counts[crit_id] = {"label": label, "count": 0}
                violation_counts[crit_id]["count"] += 1
    n = len(evaluated)
    score = round(total_points / (n * 3) * 100, 1) if n > 0 else 0.0
    grade, grade_label = "D", "Критично"
    for threshold, g, gl in _GRADE_THRESHOLDS:
        if score >= threshold:
            grade, grade_label = g, gl
            break
    top_violations = sorted(
        [{"criterion_id": k, "label": v["label"], "error_count": v["count"]} for k, v in violation_counts.items()],
        key=lambda x: x["error_count"], reverse=True,
    )[:3]
    return {"score": score, "grade": grade, "grade_label": grade_label, "top_violations": top_violations, "evaluated_count": n}


def _build_snapshot_data(instructions: list) -> dict:
    """Формирует data-словарь снимка из списка оценённых инструкций."""
    evaluated = [i for i in instructions if i.evaluation is not None]
    data = {
        "sections": [
            {
                "title": instr.title,
                "classification": instr.classification,
                "page_number": instr.page_number,
                "color": instr.evaluation.color,
                "criteria_results": instr.evaluation.criteria_results,
                "recommendations": instr.evaluation.recommendations,
                "overrides": instr.evaluation.overrides or {},
                "model_used": instr.evaluation.model_used,
                "evaluated_at": instr.evaluation.evaluated_at.isoformat()
                    if instr.evaluation.evaluated_at else None,
            }
            for instr in evaluated
        ],
        "summary": _compute_summary(evaluated),
    }
    integral = _compute_integral_from_instructions(evaluated)
    if integral:
        data["integral"] = integral
    return data


@router.get("/document/{document_id}/check-evaluated")
def check_evaluated(document_id: int, db: Session = Depends(get_db)):
    count = (
        db.query(Evaluation)
        .join(Instruction, Evaluation.instruction_id == Instruction.id)
        .filter(Instruction.document_id == document_id)
        .count()
    )
    return {"has_evaluations": count > 0, "count": count}


@router.get("/document/{document_id}/partial")
def list_partial_snapshots(document_id: int, db: Session = Depends(get_db)):
    """Возвращает список промежуточных снимков для указанного документа."""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")

    snapshots = (
        db.query(Snapshot)
        .filter(
            Snapshot.document_id == document_id,
            Snapshot.is_partial == True,
        )
        .order_by(Snapshot.created_at.desc())
        .all()
    )
    return [_serialize_snapshot(s) for s in snapshots]


@router.get("/document/{document_id}/export")
def export_xls(document_id: int, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")

    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")

    instructions = (
        db.query(Instruction)
        .filter(Instruction.document_id == document_id)
        .order_by(Instruction.id)
        .all()
    )
    evaluated = [i for i in instructions if i.evaluation is not None]
    if not evaluated:
        raise HTTPException(status_code=400, detail="Нет оценённых инструкций")

    COLOR_FILLS = {
        "green":  PatternFill("solid", fgColor="C6EFCE"),
        "yellow": PatternFill("solid", fgColor="FFEB9C"),
        "orange": PatternFill("solid", fgColor="FFCC99"),
        "red":    PatternFill("solid", fgColor="FFC7CE"),
    }
    HEADER_FILL = PatternFill("solid", fgColor="2563EB")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты оценки"

    headers = ["Раздел", "Стр.", "Оценка", "Рекомендации"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True)

    for row_idx, instr in enumerate(evaluated, 2):
        ev = instr.evaluation
        recommendations = ev.recommendations or []

        rec_text = "\n".join(
            f"[{r.get('criterion', '?')}] {r.get('text', '')}"
            + (f"\nПример: {r['example']}" if r.get('example') else "")
            for r in recommendations
        )

        ws.cell(row=row_idx, column=1, value=instr.title)
        ws.cell(row=row_idx, column=2, value=instr.page_number or "")

        color_cell = ws.cell(row=row_idx, column=3, value=COLOR_LABEL.get(ev.color, ev.color or ""))
        if ev.color in COLOR_FILLS:
            color_cell.fill = COLOR_FILLS[ev.color]

        rec_cell = ws.cell(row=row_idx, column=4, value=rec_text)
        rec_cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 80
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = doc.filename.rsplit(".", 1)[0].replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_review.xlsx"'},
    )


@router.post("/document/{document_id}", status_code=201)
def create_snapshot(document_id: int, body: CreateSnapshotRequest, db: Session = Depends(get_db)):
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")

    # Для итогового снимка группа обязательна
    if not body.is_partial:
        if not body.group_id:
            raise HTTPException(status_code=422, detail="Для итогового снимка укажите group_id")
        group = db.query(ProductGroup).filter(ProductGroup.id == body.group_id).first()
        if not group:
            raise HTTPException(status_code=404, detail="Продуктовая группа не найдена")

    instructions = (
        db.query(Instruction)
        .filter(Instruction.document_id == document_id)
        .order_by(Instruction.id)
        .all()
    )
    evaluated = [i for i in instructions if i.evaluation is not None]
    if not evaluated:
        raise HTTPException(status_code=400, detail="Нет оценённых инструкций. Сначала выполните оценку.")

    snapshot_data = _build_snapshot_data(instructions)

    # Для итогового снимка-baseline смещаем старый baseline в current
    if not body.is_partial and body.role == "baseline":
        db.query(Snapshot).filter(
            Snapshot.group_id == body.group_id,
            Snapshot.role == "baseline",
        ).update({"role": "current"}, synchronize_session=False)

    snapshot = Snapshot(
        group_id=body.group_id if not body.is_partial else None,
        document_id=document_id,
        document_filename=doc.filename,
        name=body.name or doc.filename,
        role=body.role if not body.is_partial else "partial",
        is_partial=body.is_partial,
        data=snapshot_data,
    )
    db.add(snapshot)
    db.commit()
    db.refresh(snapshot)
    return _serialize_snapshot(snapshot)


@router.post("/merge", status_code=201)
def merge_snapshots(body: MergeSnapshotsRequest, db: Session = Depends(get_db)):
    """
    Объединяет промежуточные снимки в один итоговый.

    Логика слияния: для каждого уникального заголовка раздела берётся
    результат из снимка с наибольшим created_at (последний побеждает).
    """
    if len(body.snapshot_ids) < 1:
        raise HTTPException(status_code=422, detail="Укажите хотя бы один снимок для объединения")

    # Проверяем и загружаем снимки, сортируем по дате создания (старые первыми)
    snapshots = (
        db.query(Snapshot)
        .filter(Snapshot.id.in_(body.snapshot_ids))
        .order_by(Snapshot.created_at.asc())
        .all()
    )
    if len(snapshots) != len(body.snapshot_ids):
        found_ids = {s.id for s in snapshots}
        missing = [i for i in body.snapshot_ids if i not in found_ids]
        raise HTTPException(status_code=404, detail=f"Снимки не найдены: {missing}")

    if not all(s.is_partial for s in snapshots):
        non_partial = [s.id for s in snapshots if not s.is_partial]
        raise HTTPException(
            status_code=422,
            detail=f"Снимки {non_partial} не являются промежуточными. Объединять можно только промежуточные снимки."
        )

    # Убеждаемся что все снимки относятся к одному документу
    doc_ids = {s.document_id for s in snapshots}
    if len(doc_ids) > 1:
        raise HTTPException(
            status_code=422,
            detail="Все снимки должны относиться к одному документу"
        )

    # Проверяем группу, если указана
    if body.group_id:
        group = db.query(ProductGroup).filter(ProductGroup.id == body.group_id).first()
        if not group:
            raise HTTPException(status_code=404, detail="Продуктовая группа не найдена")

    # Слияние: итерируем снимки от старых к новым, последний результат по title побеждает
    merged_by_title: dict = {}
    for snap in snapshots:
        for section in snap.data.get("sections", []):
            merged_by_title[section["title"]] = section

    merged_sections = list(merged_by_title.values())
    merged_summary = _compute_summary_from_sections(merged_sections)

    # Определяем document_filename из первого снимка
    document_filename = snapshots[0].document_filename
    document_id = snapshots[0].document_id

    # Для baseline — снимаем старый baseline
    if body.role == "baseline" and body.group_id:
        db.query(Snapshot).filter(
            Snapshot.group_id == body.group_id,
            Snapshot.role == "baseline",
        ).update({"role": "current"}, synchronize_session=False)

    merged_snapshot = Snapshot(
        group_id=body.group_id,
        document_id=document_id,
        document_filename=document_filename,
        name=body.name,
        role=body.role,
        is_partial=False,
        data={"sections": merged_sections, "summary": merged_summary},
    )
    db.add(merged_snapshot)
    db.commit()
    db.refresh(merged_snapshot)
    return _serialize_snapshot(merged_snapshot)


@router.delete("/{snapshot_id}", status_code=204)
def delete_snapshot(snapshot_id: int, db: Session = Depends(get_db)):
    snapshot = db.query(Snapshot).filter(Snapshot.id == snapshot_id).first()
    if not snapshot:
        raise HTTPException(status_code=404, detail="Снимок не найден")
    db.delete(snapshot)
    db.commit()


@router.get("/compare")
def compare_snapshots(
    document_id: Optional[int] = Query(None),
    snapshot_a: int = Query(...),
    snapshot_b: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    snap_a = db.query(Snapshot).filter(Snapshot.id == snapshot_a).first()
    if not snap_a:
        raise HTTPException(status_code=404, detail=f"Снимок A (id={snapshot_a}) не найден")

    sections_a = snap_a.data["sections"]
    source_a = {
        "type": "snapshot", "id": snap_a.id, "name": snap_a.name,
        "role": snap_a.role, "document_filename": snap_a.document_filename,
        "created_at": snap_a.created_at.isoformat(),
    }

    if snapshot_b:
        snap_b = db.query(Snapshot).filter(Snapshot.id == snapshot_b).first()
        if not snap_b:
            raise HTTPException(status_code=404, detail=f"Снимок B (id={snapshot_b}) не найден")
        sections_b = snap_b.data["sections"]
        source_b = {
            "type": "snapshot", "id": snap_b.id, "name": snap_b.name,
            "role": snap_b.role, "document_filename": snap_b.document_filename,
            "created_at": snap_b.created_at.isoformat(),
        }
    else:
        if not document_id:
            raise HTTPException(status_code=422, detail="Укажите document_id для сравнения с текущим состоянием")
        instructions = (
            db.query(Instruction)
            .filter(Instruction.document_id == document_id)
            .order_by(Instruction.id)
            .all()
        )
        evaluated = [i for i in instructions if i.evaluation is not None]
        evaluated = [i for i in instructions if i.evaluation is not None]
        if not evaluated:
            raise HTTPException(status_code=400, detail="Нет оценённых инструкций для сравнения.")
        doc = db.query(Document).filter(Document.id == document_id).first()
        sections_b = [
            {"title": i.title, "classification": i.classification, "page_number": i.page_number,
             "color": i.evaluation.color, "criteria_results": i.evaluation.criteria_results,
             "overrides": i.evaluation.overrides or {}}
            for i in evaluated
        ]
        source_b = {
            "type": "current", "document_id": document_id,
            "document_filename": doc.filename if doc else None,
        }

    # Строим diff
    map_a = {s["title"]: s for s in sections_a}
    map_b = {s["title"]: s for s in sections_b}
    all_titles = list(dict.fromkeys(list(map_a.keys()) + list(map_b.keys())))

    COLOR_ORDER = ["green", "yellow", "orange", "red"]

    def _change(ca, cb):
        if ca is None: return "new"
        if cb is None: return "removed"
        ia = COLOR_ORDER.index(ca) if ca in COLOR_ORDER else -1
        ib = COLOR_ORDER.index(cb) if cb in COLOR_ORDER else -1
        if ia == ib: return "unchanged"
        return "improved" if ib < ia else "degraded"

    diff = [
        {
            "title": t,
            "color_a": map_a.get(t, {}).get("color"),
            "color_b": map_b.get(t, {}).get("color"),
            "page_number": (map_b.get(t) or map_a.get(t) or {}).get("page_number"),
            "change": _change(
                map_a.get(t, {}).get("color"),
                map_b.get(t, {}).get("color"),
            ),
        }
        for t in all_titles
    ]

    stats = {"degraded": 0, "improved": 0, "new": 0, "removed": 0, "unchanged": 0}
    for row in diff:
        stats[row["change"]] += 1

    return {"source_a": source_a, "source_b": source_b, "diff": diff, "stats": stats}
